from __future__ import annotations

import asyncio
import json
import re
import tempfile
import zipfile
from collections import defaultdict
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import asyncpg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

from gaokao_vault.db.queries.liaoning import (
    LIAONING_PROVINCE,
    fetch_liaoning_export_admissions,
    fetch_liaoning_export_plans,
    fetch_liaoning_historical_plans,
    fetch_liaoning_school_charters,
)
from gaokao_vault.pipeline.batch_normalizer import BatchInfo, normalize_batch

PLAN_HEADERS = (
    "ID",
    "年份",
    "生源地",
    "批次",
    "科类",
    "计划类别",
    "院校代码",
    "院校名称",
    "专业代码",
    "专业全称",
    "专业名称",
    "专业备注",
    "选科要求",
    "专业层次",
    "计划人数",
    "学制",
    "学费",
    "门类",
    "专业类",
    "是否新增",
)
HISTORY_HEADERS = ("录取人数", "最低分", "最低位次", "老批次", "计划人数结果")
SCHOOL_HEADERS = (
    "所在省",
    "城市",
    "城市水平标签",
    "院校标签",
    "院校水平",
    "更名合并转设",
    "隶属单位",
    "类型",
    "公私性质",
    "本科/专科",
    "保研率",
    "院校排名",
    "转专业情况",
    "全校硕士专业数",
    "全校硕士专业",
    "全校博士专业数",
    "全校博士专业",
    "录取规则",
    "招生章程",
)
MAJOR_HEADERS = ("软科评级", "软科排名", "学科评估", "专业水平", "本专业硕士点", "本专业博士点")

_MAJOR_BRACKET_RE = re.compile(r"[\uff08(].*?[\uff09)]")
_DURATION_RE = re.compile(r"\d+")
_LEFT_ALIGNED_COLUMNS = frozenset({8, 10, 11, 12, 13, 18, 19, *range(41, 66)})
_TEXT_COLUMNS = frozenset({1, 7, 9})
_STYLES_PART = "xl/styles.xml"
_SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_FONT_ELEMENT_ORDER = {
    name: index
    for index, name in enumerate((
        "b",
        "i",
        "strike",
        "outline",
        "shadow",
        "condense",
        "extend",
        "u",
        "vertAlign",
        "sz",
        "color",
        "name",
        "family",
        "charset",
        "scheme",
    ))
}


@dataclass(frozen=True, slots=True)
class ExportSummary:
    output_path: Path
    row_count: int
    history_years: tuple[int, ...]
    matched_history_counts: dict[int, int]
    new_plan_count: int


def history_years_for(plan_year: int) -> tuple[int, int, int, int]:
    return plan_year - 1, plan_year - 2, plan_year - 3, plan_year - 4


async def export_liaoning_workbook(
    pool: asyncpg.Pool,
    output_path: Path,
    *,
    plan_year: int,
    subject: str | None = None,
    baseline_path: Path | None = None,
) -> ExportSummary:
    history_years = history_years_for(plan_year)
    async with pool.acquire() as conn:
        plans = await fetch_liaoning_export_plans(conn, plan_year=plan_year, subject=subject)
        if not plans:
            msg = f"数据库中没有辽宁 {plan_year} 年招生计划数据"
            raise ValueError(msg)
        admissions = await fetch_liaoning_export_admissions(conn, years=history_years, subject=subject)
        historical_plans = await fetch_liaoning_historical_plans(conn, years=history_years, subject=subject)
        charters = await fetch_liaoning_school_charters(conn, plan_year=plan_year)

    rows, matched_counts, new_count = build_liaoning_export_rows(
        plans,
        admissions,
        historical_plans,
        charters,
        plan_year=plan_year,
    )
    if baseline_path is not None:
        baseline_rows = await asyncio.to_thread(load_liaoning_baseline_rows, baseline_path)
        rows, matched_counts, new_count = merge_liaoning_baseline_rows(
            rows,
            baseline_rows,
            plan_year=plan_year,
        )
    await asyncio.to_thread(write_liaoning_workbook, output_path, rows, plan_year=plan_year)
    return ExportSummary(
        output_path=output_path,
        row_count=len(rows),
        history_years=history_years,
        matched_history_counts=matched_counts,
        new_plan_count=new_count,
    )


def build_liaoning_export_rows(
    plans: Sequence[dict[str, Any]],
    admissions: Sequence[dict[str, Any]],
    historical_plans: Sequence[dict[str, Any]],
    charters: Sequence[dict[str, Any]],
    *,
    plan_year: int,
) -> tuple[list[list[Any]], dict[int, int], int]:
    history_years = history_years_for(plan_year)
    admission_indexes = _build_history_indexes(admissions)
    historical_plan_indexes = _build_history_indexes(historical_plans)
    charter_by_school = {int(row["school_id"]): row for row in charters}
    matched_counts = dict.fromkeys(history_years, 0)
    new_count = 0
    output_rows: list[list[Any]] = []

    for plan in plans:
        matches: dict[int, dict[str, Any] | None] = {}
        historical_plan_matches: dict[int, dict[str, Any] | None] = {}
        for year in history_years:
            match = _match_history_row(plan, year, admission_indexes)
            matches[year] = match
            historical_plan_matches[year] = _match_history_row(plan, year, historical_plan_indexes)
            if match is not None:
                matched_counts[year] += 1

        is_new = all(
            matches[year] is None and historical_plan_matches[year] is None
            for year in history_years
        )
        if is_new:
            new_count += 1

        canonical_major_name = _clean_text(plan.get("canonical_major_name")) or _base_major_name(
            plan.get("major_full_name")
        )
        major_full_name = _clean_text(plan.get("major_full_name")) or canonical_major_name
        charter = charter_by_school.get(int(plan["school_id"]))
        row: list[Any] = [
            _as_text(plan.get("plan_id")),
            plan_year,
            LIAONING_PROVINCE,
            plan.get("batch"),
            _subject_label(plan.get("subject_category")),
            _plan_category(plan.get("program_type")),
            _as_text(plan.get("school_code_raw")),
            plan.get("school_name"),
            _as_text(plan.get("major_code_raw")),
            major_full_name,
            canonical_major_name,
            _major_note(major_full_name, canonical_major_name, plan.get("note")),
            _selection_requirement(plan.get("selection_requirement")),
            _education_level(plan.get("education_level"), plan.get("batch")),
            plan.get("plan_count"),
            _duration(plan.get("duration")),
            plan.get("tuition"),
            plan.get("major_category"),
            plan.get("major_subcategory"),
            "新增" if is_new else None,
        ]

        for year in history_years:
            admission = matches[year]
            historical_plan = historical_plan_matches[year]
            historical_plan_count = admission.get("plan_count") if admission else None
            if historical_plan_count is None and historical_plan is not None:
                historical_plan_count = historical_plan.get("plan_count")
            row.extend([
                admission.get("admitted_count") if admission else None,
                admission.get("min_score") if admission else None,
                admission.get("min_rank") if admission else None,
                admission.get("batch") if admission else None,
                historical_plan_count,
            ])

        row.extend([
            plan.get("school_province"),
            plan.get("city"),
            None,
            _school_tags(plan),
            None,
            None,
            plan.get("authority"),
            plan.get("school_type"),
            _ownership_type(plan),
            plan.get("school_level"),
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            _admission_rule(plan, charter),
            _charter_url(plan, charter),
            None,
            None,
            None,
            _strength_label(plan.get("strength_evidence"), plan.get("major_strength_tier")),
            None,
            None,
        ])
        if len(row) != 65:
            msg = f"辽宁导出行列数异常: expected=65 actual={len(row)} plan_id={plan.get('plan_id')}"
            raise ValueError(msg)
        output_rows.append(row)

    return output_rows, matched_counts, new_count


def load_liaoning_baseline_rows(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows: list[list[Any]] = []
    try:
        for values in worksheet.iter_rows(min_row=3, values_only=True):
            row = list(values[:65])
            if not any(value not in (None, "") for value in row):
                continue
            row.extend([None] * (65 - len(row)))
            rows.append(row)
    finally:
        workbook.close()
    return rows


def merge_liaoning_baseline_rows(
    generated_rows: Sequence[list[Any]],
    baseline_rows: Sequence[list[Any]],
    *,
    plan_year: int,
) -> tuple[list[list[Any]], dict[int, int], int]:
    exact_index: dict[tuple[str, ...], list[int]] = defaultdict(list)
    name_index: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for index, row in enumerate(baseline_rows):
        exact_index[_export_row_key(row)].append(index)
        name_index[_export_row_name_key(row)].append(index)

    merged = [list(row) for row in baseline_rows]
    consumed: set[int] = set()
    for generated in generated_rows:
        candidates = [index for index in exact_index.get(_export_row_key(generated), []) if index not in consumed]
        if len(candidates) != 1:
            candidates = [
                index for index in name_index.get(_export_row_name_key(generated), []) if index not in consumed
            ]
        if len(candidates) == 1:
            index = candidates[0]
            consumed.add(index)
            merged[index] = [*generated[:20], *merged[index][20:65]]
        else:
            merged.append(list(generated))

    history_years = history_years_for(plan_year)
    matched_counts = dict.fromkeys(history_years, 0)
    new_count = 0
    for row in merged:
        has_history = False
        for offset, year in enumerate(history_years):
            block = row[20 + offset * 5 : 25 + offset * 5]
            if any(value not in (None, "") for value in block):
                matched_counts[year] += 1
                has_history = True
        row[19] = None if has_history else "新增"
        new_count += not has_history
    return merged, matched_counts, new_count


def _export_row_key(row: Sequence[Any]) -> tuple[str, ...]:
    return (
        _normalized_code(row[6]),
        _clean_text(row[7]),
        _normalized_code(row[8]),
        _normalized_full_major_name(row[9]),
        _subject_label(row[4]),
        _normalized_batch_text(row[3]),
    )


def _export_row_name_key(row: Sequence[Any]) -> tuple[str, ...]:
    return (
        _clean_text(row[7]),
        _normalized_full_major_name(row[9]),
        _subject_label(row[4]),
        _normalized_batch_text(row[3]),
    )


def _build_history_indexes(
    rows: Sequence[dict[str, Any]],
) -> tuple[dict[tuple, list[dict[str, Any]]], dict[tuple, list[dict[str, Any]]]]:
    by_id: dict[tuple, list[dict[str, Any]]] = defaultdict(list)
    by_name: dict[tuple, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        subject = _subject_label(row.get("subject_category"))
        year = int(row["year"])
        school_id = int(row["school_id"])
        major_id = row.get("major_id")
        if major_id is not None:
            by_id[(year, school_id, int(major_id), subject)].append(row)
        major_name = _normalized_major_name(
            row.get("canonical_major_name") or row.get("major_name_raw") or row.get("major_name")
        )
        if major_name:
            by_name[(year, school_id, major_name, subject)].append(row)
    return by_id, by_name


def _match_history_row(
    plan: dict[str, Any],
    year: int,
    indexes: tuple[dict[tuple, list[dict[str, Any]]], dict[tuple, list[dict[str, Any]]]],
) -> dict[str, Any] | None:
    by_id, by_name = indexes
    subject = _subject_label(plan.get("subject_category"))
    school_id = int(plan["school_id"])
    candidates: list[dict[str, Any]] = []
    major_id = plan.get("major_id")
    if major_id is not None:
        candidates = by_id.get((year, school_id, int(major_id), subject), [])
        if not candidates:
            return None
    else:
        major_name = _normalized_major_name(plan.get("canonical_major_name") or plan.get("major_full_name"))
        candidates = by_name.get((year, school_id, major_name, subject), [])
        candidate_major_ids = {
            candidate.get("major_id") for candidate in candidates if candidate.get("major_id") is not None
        }
        if len(candidate_major_ids) > 1:
            return None

    candidates = [candidate for candidate in candidates if _batch_match_score(plan, candidate) >= 0]
    if not candidates:
        return None

    best_batch_score = max(_batch_match_score(plan, candidate) for candidate in candidates)
    candidates = [candidate for candidate in candidates if _batch_match_score(plan, candidate) == best_batch_score]
    candidates = _prefer_exact_variant(plan, candidates)
    if _has_ambiguous_variants(candidates):
        return None
    return max(candidates, key=lambda candidate: (_row_completeness(candidate), _source_priority(candidate)))


def _prefer_exact_variant(
    plan: dict[str, Any],
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    plan_school_code = _normalized_code(plan.get("school_code_raw"))
    plan_major_code = _normalized_code(plan.get("major_code_raw"))
    plan_group_code = _normalized_code(plan.get("major_group_code"))
    plan_full_name = _normalized_full_major_name(plan.get("major_full_name") or plan.get("major_name"))

    selectors = (
        lambda candidate: bool(
            plan_school_code
            and plan_major_code
            and _normalized_code(candidate.get("school_code_raw")) == plan_school_code
            and _normalized_code(candidate.get("major_code_raw")) == plan_major_code
        ),
        lambda candidate: bool(
            plan_group_code
            and plan_major_code
            and _normalized_code(candidate.get("major_group_code")) == plan_group_code
            and _normalized_code(candidate.get("major_code_raw")) == plan_major_code
        ),
        lambda candidate: bool(
            plan_major_code and _normalized_code(candidate.get("major_code_raw")) == plan_major_code
        ),
        lambda candidate: bool(
            plan_full_name
            and _normalized_full_major_name(
                candidate.get("major_name_raw") or candidate.get("major_name") or candidate.get("canonical_major_name")
            )
            == plan_full_name
        ),
    )
    for selector in selectors:
        exact = [candidate for candidate in candidates if selector(candidate)]
        if exact:
            return exact
    return candidates


def _has_ambiguous_variants(candidates: Sequence[dict[str, Any]]) -> bool:
    if len(candidates) <= 1:
        return False
    identities = {
        (
            _normalized_code(candidate.get("school_code_raw")),
            _normalized_code(candidate.get("major_group_code")),
            _normalized_code(candidate.get("major_code_raw")),
            _normalized_full_major_name(
                candidate.get("major_name_raw") or candidate.get("major_name") or candidate.get("canonical_major_name")
            ),
        )
        for candidate in candidates
    }
    return len(identities) > 1


def _source_priority(row: dict[str, Any]) -> int:
    return 1 if row.get("data_source") == "gaokao.chsi.com.cn" else 0


def _batch_match_score(left: dict[str, Any], right: dict[str, Any]) -> int:
    left_info = _batch_info(left)
    right_info = _batch_info(right)
    if left_info.code and right_info.code:
        if left_info.code != right_info.code:
            return -1
        if left_info.segment and right_info.segment and left_info.segment != right_info.segment:
            return -1
        return 4 if left_info.segment == right_info.segment else 3

    left_raw = _normalized_batch_text(left.get("batch"))
    right_raw = _normalized_batch_text(right.get("batch"))
    if left_raw and right_raw and left_raw == right_raw:
        return 2
    return 0 if not left_raw or not right_raw else -1


def _batch_info(row: dict[str, Any]) -> BatchInfo:
    code = _clean_text(row.get("batch_code"))
    category = _clean_text(row.get("batch_category"))
    segment = _clean_text(row.get("batch_segment"))
    if code or category or segment:
        return BatchInfo(code=code, category=category, segment=segment)
    return normalize_batch(_clean_text(row.get("batch")))


def _row_completeness(row: dict[str, Any]) -> int:
    return sum(
        value not in (None, "")
        for value in (
            row.get("min_rank"),
            row.get("min_score"),
            row.get("admitted_count"),
            row.get("plan_count"),
        )
    )


def _normalized_major_name(value: Any) -> str:
    text = _MAJOR_BRACKET_RE.sub("", _clean_text(value)).lower()
    return re.sub(r"\s+|专业|类", "", text)


def _normalized_full_major_name(value: Any) -> str:
    text = _clean_text(value).lower().translate(str.maketrans({"\uff08": "(", "\uff09": ")", "\uff0c": ","}))
    return re.sub(r"\s+", "", text)


def _normalized_code(value: Any) -> str:
    return re.sub(r"\s+", "", _clean_text(value)).upper()


def _normalized_batch_text(value: Any) -> str:
    return re.sub(
        r"普通类|院校最低分|综合评价分数|[\u2014\-\uff0d()\uff08\uff09\s]",
        "",
        _clean_text(value),
    )


def _subject_label(value: Any) -> str:
    return _clean_text(value).removesuffix("类")


def _base_major_name(value: Any) -> str | None:
    text = _clean_text(value)
    base = re.split(r"[\uff08(]", text, maxsplit=1)[0].strip()
    return base or None


def _major_note(full_name: Any, canonical_name: Any, note: Any) -> str | None:
    full = _clean_text(full_name)
    canonical = _clean_text(canonical_name)
    suffix = full[len(canonical) :].strip() if canonical and full.startswith(canonical) else ""
    values = [value for value in (suffix, _clean_text(note)) if value]
    return "\uff1b".join(dict.fromkeys(values)) or None


def _selection_requirement(value: Any) -> str | None:
    text = _clean_text(value)
    for prefix in ("首选物理,再选", "首选历史,再选", "物理+", "历史+"):
        if text.startswith(prefix):
            text = text.removeprefix(prefix).strip()
    return text or None


def _education_level(value: Any, batch: Any) -> str | None:
    text = _clean_text(value)
    if "职业" in text and "本科" in text:
        return "职教本科"
    if text:
        return text
    batch_text = _clean_text(batch)
    if "专科" in batch_text or "高职" in batch_text:
        return "专科"
    if "本科" in batch_text:
        return "本科"
    return None


def _duration(value: Any) -> str | None:
    text = _clean_text(value)
    match = _DURATION_RE.search(text)
    if match:
        return match.group(0)
    chinese_years = {"一": "1", "二": "2", "三": "3", "四": "4", "五": "5", "六": "6", "七": "7", "八": "8"}
    for chinese, number in chinese_years.items():
        if chinese in text:
            return number
    return text or None


def _plan_category(value: Any) -> str | None:
    text = _clean_text(value)
    if not text or text in {"普通", "普通类"}:
        return None
    return text if text.startswith("(") else f"({text})"


def _school_tags(plan: dict[str, Any]) -> str | None:
    tags = []
    if plan.get("is_985"):
        tags.append("985")
    if plan.get("is_211"):
        tags.append("211")
    if plan.get("is_double_first"):
        tags.append("双一流")
    return "/".join(tags) or None


def _ownership_type(plan: dict[str, Any]) -> str:
    if plan.get("is_sino_foreign"):
        return "中外合作办学"
    if plan.get("is_private") or plan.get("is_independent"):
        return "民办"
    return "公办"


def _strength_label(evidence: Any, tier: Any) -> str | None:
    if isinstance(evidence, str):
        try:
            evidence = json.loads(evidence)
        except json.JSONDecodeError:
            evidence = []
    labels: list[str] = []
    for item in evidence or []:
        if not isinstance(item, dict):
            continue
        signal_type = item.get("signal_type")
        signal_level = item.get("signal_level")
        if signal_type == "first_class_major" and signal_level == "national":
            labels.append("国一")
        elif signal_type == "first_class_major" and signal_level == "provincial":
            labels.append("省一")
        elif signal_type == "featured_major" and signal_level == "national":
            labels.append("国特")
    if not labels and tier == "national":
        labels.append("国一")
    elif not labels and tier == "provincial":
        labels.append("省一")
    return "/".join(dict.fromkeys(labels)) or None


def _admission_rule(plan: dict[str, Any], charter: dict[str, Any] | None) -> str | None:
    content = _clean_text(charter.get("content")) if charter else ""
    if "专业志愿优先" in content or "专业优先" in content:
        return "专业优先"
    if "分数优先" in content:
        return "分数优先"
    if "专业级差" in content:
        return "专业级差"
    return _clean_text(plan.get("adjustment_rule")) or None


def _charter_url(plan: dict[str, Any], charter: dict[str, Any] | None) -> str | None:
    if charter and charter.get("source_url"):
        return str(charter["source_url"])
    sch_id = plan.get("sch_id")
    if sch_id is None:
        return None
    return f"https://gaokao.chsi.com.cn/zsgs/zhangcheng/listZszc--schId-{sch_id}.dhtml"


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _as_text(value: Any) -> str | None:
    text = _clean_text(value)
    return text or None


def write_liaoning_workbook(
    output_path: Path,
    rows: Iterable[Sequence[Any]],
    *,
    plan_year: int,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    history_years = history_years_for(plan_year)

    group_row: list[Any] = [None] * 65
    group_row[0] = f"{plan_year}年招生计划"
    for index, year in enumerate(history_years):
        group_row[20 + index * 5] = f"{year}年专业录取数据"
    group_row[40] = "院校基础信息"
    group_row[59] = "专业基础信息"
    worksheet.append(group_row)
    worksheet.append(_column_headers())

    merge_ranges = ("A1:S1", "U1:Y1", "Z1:AD1", "AE1:AI1", "AJ1:AN1", "AO1:BG1", "BH1:BM1")
    for merge_range in merge_ranges:
        worksheet.merge_cells(merge_range)

    _register_styles(workbook)
    _style_headers(worksheet)

    for values in rows:
        worksheet.append(list(values))
        row_index = worksheet.max_row
        for column_index in range(1, 66):
            cell = worksheet.cell(row_index, column_index)
            cell.style = "data_left" if column_index in _LEFT_ALIGNED_COLUMNS else "data_center"
            if column_index in _TEXT_COLUMNS:
                cell.number_format = "@"
        if worksheet.cell(row_index, 59).value:
            worksheet.cell(row_index, 59).hyperlink = str(worksheet.cell(row_index, 59).value)
            worksheet.cell(row_index, 59).style = "hyperlink"

    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:BM{max(2, worksheet.max_row)}"
    worksheet.sheet_view.showGridLines = False
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = "1:2"
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 41.65
    _set_column_widths(worksheet)

    with tempfile.NamedTemporaryFile(dir=output_path.parent, suffix=".xlsx", delete=False) as temp_file:
        temp_path = Path(temp_file.name)
    try:
        workbook.save(temp_path)
        _normalize_openxml_styles(temp_path)
        temp_path.replace(output_path)
    finally:
        temp_path.unlink(missing_ok=True)


def _normalize_openxml_styles(workbook_path: Path) -> None:
    with tempfile.NamedTemporaryFile(dir=workbook_path.parent, suffix=".xlsx", delete=False) as temp_file:
        normalized_path = Path(temp_file.name)
    try:
        with (
            zipfile.ZipFile(workbook_path, "r") as source,
            zipfile.ZipFile(
                normalized_path,
                "w",
            ) as target,
        ):
            for info in source.infolist():
                content = source.read(info.filename)
                if info.filename == _STYLES_PART:
                    content = _normalized_styles_xml(content)
                target.writestr(info, content)
        normalized_path.replace(workbook_path)
    finally:
        normalized_path.unlink(missing_ok=True)


def _normalized_styles_xml(content: bytes) -> bytes:
    root = ElementTree.fromstring(content)  # noqa: S314 - 只解析本进程刚生成的 OOXML
    font_tag = f"{{{_SPREADSHEET_NS}}}font"
    for font in root.iter(font_tag):
        children = list(font)
        children.sort(key=lambda child: _FONT_ELEMENT_ORDER.get(child.tag.rsplit("}", 1)[-1], len(_FONT_ELEMENT_ORDER)))
        font[:] = children
    ElementTree.register_namespace("", _SPREADSHEET_NS)
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _column_headers() -> list[str]:
    headers = list(PLAN_HEADERS)
    for index in range(1, 5):
        headers.extend(f"{header}{index}" for header in HISTORY_HEADERS)
    headers.extend(SCHOOL_HEADERS)
    headers.extend(MAJOR_HEADERS)
    return headers


def _register_styles(workbook: Workbook) -> None:
    data_center = NamedStyle(name="data_center")
    data_center.font = Font(name="Arial", size=10)
    data_center.alignment = Alignment(horizontal="center", vertical="center")
    workbook.add_named_style(data_center)

    data_left = NamedStyle(name="data_left")
    data_left.font = Font(name="Arial", size=10)
    data_left.alignment = Alignment(horizontal="left", vertical="center")
    workbook.add_named_style(data_left)

    hyperlink = NamedStyle(name="hyperlink")
    hyperlink.font = Font(name="Arial", size=10, color="0563C1", underline="single")
    hyperlink.alignment = Alignment(horizontal="left", vertical="center")
    workbook.add_named_style(hyperlink)


def _style_headers(worksheet) -> None:
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "plan": PatternFill("solid", fgColor="DDEBF7"),
        "history": PatternFill("solid", fgColor="FFF2CC"),
        "school": PatternFill("solid", fgColor="FCE4D6"),
        "major": PatternFill("solid", fgColor="E2EFDA"),
    }
    for row in (1, 2):
        for column in range(1, 66):
            cell = worksheet.cell(row, column)
            cell.fill = fills[_column_group(column)]
            cell.font = Font(name="Arial", size=11, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    for row in (2,):
        worksheet.cell(row, 12).fill = PatternFill("solid", fgColor="00B050")
        worksheet.cell(row, 15).fill = PatternFill("solid", fgColor="FFFF00")
        worksheet.cell(row, 21).fill = PatternFill("solid", fgColor="FFFF00")


def _column_group(column: int) -> str:
    if column <= 19:
        return "plan"
    if column <= 40:
        return "history"
    if column <= 59:
        return "school"
    return "major"


def _set_column_widths(worksheet) -> None:
    widths = [
        8,
        8,
        9,
        22,
        9,
        18,
        11,
        22,
        11,
        44,
        22,
        46,
        18,
        12,
        10,
        8,
        12,
        15,
        18,
        10,
        *([11, 11, 13, 22, 13] * 4),
        10,
        14,
        16,
        24,
        28,
        32,
        16,
        14,
        16,
        12,
        10,
        14,
        34,
        12,
        34,
        12,
        34,
        16,
        34,
        12,
        12,
        16,
        20,
        24,
        24,
    ]
    if len(widths) != 65:
        msg = f"辽宁工作簿列宽配置异常: expected=65 actual={len(widths)}"
        raise ValueError(msg)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
