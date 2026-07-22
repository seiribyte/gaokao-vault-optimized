from __future__ import annotations

import asyncio
import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast
from urllib.parse import quote
from urllib.request import Request as UrlRequest
from urllib.request import urlopen

import asyncpg
from openpyxl import load_workbook

_SEARCH_URL = "https://gaokao.chsi.com.cn/sch/search.do?searchType=1&yxmc={}"
_INFO_URL = "https://gaokao.chsi.com.cn/wap/sch/schinfo/{}"
_GAOKAO_INDEX_URL = "https://static-data.gaokao.cn/www/2.0/school/name.json"
_SCHOOL_ID_RE = re.compile(r"schoolInfo--schId-(\d+)")
_REFERENCE_ALIASES = {
    "中国人民大学(苏州校区)": "中国人民大学",
    "北京交通大学(威海校区)": "北京交通大学",
    "西南大学(荣昌校区)": "西南大学",
}


@dataclass(slots=True)
class _GaokaoCatalogState:
    existing_by_name: dict[str, Any]
    existing_by_id: dict[int, Any]
    occupied_ids: set[int]
    occupied_gaokao_ids: dict[int, int]


async def sync_reference_schools(conn: asyncpg.Connection, reference_path: Path) -> dict[str, int]:
    targets = await asyncio.to_thread(_read_reference_schools, reference_path)
    existing_rows = await conn.fetch("SELECT id, name, sch_id FROM schools")
    existing_by_name: dict[str, Any] = {}
    ordered_rows = sorted(
        existing_rows,
        key=lambda item: (int(item["sch_id"]) > 0, -int(item["id"])),
        reverse=True,
    )
    for row in ordered_rows:
        if row["name"]:
            existing_by_name.setdefault(str(row["name"]).strip(), row)
    pending = [
        target
        for target in targets
        if target["name"] not in existing_by_name or int(existing_by_name[target["name"]]["sch_id"]) < 0
    ]
    results = await asyncio.gather(*(asyncio.to_thread(_resolve_school, target) for target in pending))
    resolved = [result for result in results if result is not None]
    province_rows = await conn.fetch("SELECT id, name FROM provinces")
    province_ids = {_normalize_province_name(str(row["name"])): int(row["id"]) for row in province_rows}
    occupied_ids = {int(row["sch_id"]): int(row["id"]) for row in existing_rows}
    async with conn.transaction():
        for school in resolved:
            school["province_id"] = province_ids.get(_normalize_province_name(str(school.get("province") or "")))
            name = str(school["name"])
            existing = existing_by_name.get(name)
            sch_id = _safe_catalog_sch_id(school)
            owner_id = occupied_ids.get(sch_id)
            if owner_id is not None and (existing is None or owner_id != int(existing["id"])):
                sch_id = (
                    int(existing["sch_id"])
                    if existing is not None and int(existing["sch_id"]) < 0
                    else _allocate_catalog_sch_id(name, set(occupied_ids))
                )

            if existing is not None and int(existing["sch_id"]) < 0:
                old_sch_id = int(existing["sch_id"])
                await conn.execute(
                    """
                    UPDATE schools
                    SET sch_id = $2,
                        province_id = COALESCE($3, province_id),
                        city = COALESCE($4, city),
                        authority = COALESCE($5, authority),
                        level = COALESCE($6, level),
                        is_double_first = $7
                    WHERE id = $1
                    """,
                    int(existing["id"]),
                    sch_id,
                    school["province_id"],
                    school["city"],
                    school["authority"],
                    school["level"],
                    school["is_double_first"],
                )
                if old_sch_id != sch_id:
                    occupied_ids.pop(old_sch_id, None)
                occupied_ids[sch_id] = int(existing["id"])
                existing_by_name[name] = {**dict(existing), "sch_id": sch_id}
                continue
            await conn.execute(
                """
                INSERT INTO schools (sch_id, name, province_id, city, authority, level,
                    is_211, is_985, is_double_first, is_private, is_independent, is_sino_foreign,
                    content_hash, crawl_task_id)
                VALUES ($1, $2, $3, $4, $5, $6, FALSE, FALSE, $7, FALSE, FALSE, FALSE, NULL, NULL)
                ON CONFLICT (sch_id) DO NOTHING
                """,
                sch_id,
                name,
                school["province_id"],
                school["city"],
                school["authority"],
                school["level"],
                school["is_double_first"],
            )
            occupied_ids[sch_id] = -1
    return {
        "reference_schools": len(targets),
        "already_present": len(targets) - len(pending),
        "resolved": len(resolved),
        "unresolved": len(pending) - len(resolved),
    }


async def sync_gaokao_school_index(conn: asyncpg.Connection) -> dict[str, int]:
    payload = await asyncio.to_thread(_fetch_json, _GAOKAO_INDEX_URL)
    rows = payload.get("data") if isinstance(payload, dict) else None
    targets = [row for row in rows if isinstance(row, dict) and row.get("name")] if isinstance(rows, list) else []
    existing_rows = await conn.fetch(
        """
        SELECT id, name, sch_id, gaokao_school_id, province_id, city, authority, level,
               is_211, is_985, is_double_first, is_private, is_independent, is_sino_foreign,
               school_type, website, phone, email, address, introduction, logo_url,
               content_hash, crawl_task_id
        FROM schools
        """
    )
    state = _build_gaokao_catalog_state(existing_rows)
    valid_targets, counts = _prepare_gaokao_targets(targets)
    async with conn.transaction():
        for name, gaokao_school_id in valid_targets:
            outcome = await _sync_gaokao_school(conn, state, name, gaokao_school_id)
            counts[outcome] += 1
    return {
        "index_schools": len(targets),
        "already_present": counts["already_present"],
        "linked": counts["linked"] + counts["migrated"],
        "added": counts["added"],
        "migrated": counts["migrated"],
        "conflicts": counts["conflicts"],
        "invalid": counts["invalid"],
        "duplicates": counts["duplicates"],
    }


def _build_gaokao_catalog_state(existing_rows: list[Any]) -> _GaokaoCatalogState:
    existing_by_name: dict[str, Any] = {}
    for row in sorted(existing_rows, key=_school_row_priority, reverse=True):
        if row["name"]:
            existing_by_name.setdefault(str(row["name"]).strip(), row)
    return _GaokaoCatalogState(
        existing_by_name=existing_by_name,
        existing_by_id={int(row["id"]): row for row in existing_rows},
        occupied_ids={int(row["sch_id"]) for row in existing_rows},
        occupied_gaokao_ids={
            int(row["gaokao_school_id"]): int(row["id"]) for row in existing_rows if row["gaokao_school_id"] is not None
        },
    )


def _prepare_gaokao_targets(targets: list[dict[str, Any]]) -> tuple[list[tuple[str, int]], Counter[str]]:
    valid_targets: list[tuple[str, int]] = []
    counts: Counter[str] = Counter()
    seen_names: dict[str, int] = {}
    seen_ids: dict[int, str] = {}
    for row in targets:
        name = str(row["name"]).strip()
        raw_school_id = row.get("school_id")
        if raw_school_id is None:
            counts["invalid"] += 1
            continue
        try:
            gaokao_school_id = int(cast(int | str, raw_school_id))
        except (TypeError, ValueError):
            counts["invalid"] += 1
            continue
        if not name or gaokao_school_id <= 0:
            counts["invalid"] += 1
            continue
        if name in seen_names or gaokao_school_id in seen_ids:
            if seen_names.get(name) == gaokao_school_id and seen_ids.get(gaokao_school_id) == name:
                counts["duplicates"] += 1
            else:
                counts["conflicts"] += 1
            continue
        seen_names[name] = gaokao_school_id
        seen_ids[gaokao_school_id] = name
        valid_targets.append((name, gaokao_school_id))
    return valid_targets, counts


async def _sync_gaokao_school(
    conn: asyncpg.Connection,
    state: _GaokaoCatalogState,
    name: str,
    gaokao_school_id: int,
) -> str:
    existing = state.existing_by_name.get(name)
    if existing is None:
        return await _insert_gaokao_placeholder(conn, state, name, gaokao_school_id)
    return await _link_gaokao_school(conn, state, existing, name, gaokao_school_id)


async def _link_gaokao_school(
    conn: asyncpg.Connection,
    state: _GaokaoCatalogState,
    existing: Any,
    name: str,
    gaokao_school_id: int,
) -> str:
    existing_id = int(existing["id"])
    existing_gaokao_id = existing.get("gaokao_school_id")
    if existing_gaokao_id is not None:
        return "already_present" if int(existing_gaokao_id) == gaokao_school_id else "conflicts"

    owner_id = state.occupied_gaokao_ids.get(gaokao_school_id)
    if owner_id is not None and owner_id != existing_id:
        owner = state.existing_by_id.get(owner_id)
        if owner is None or str(owner["name"]).strip() != name:
            return "conflicts"
        await conn.execute("UPDATE schools SET gaokao_school_id = NULL WHERE id = $1", owner_id)
        state.occupied_gaokao_ids.pop(gaokao_school_id, None)

    legacy_placeholder = _is_legacy_gaokao_placeholder(existing, gaokao_school_id)
    sch_id = int(existing["sch_id"])
    if legacy_placeholder:
        new_sch_id = _allocate_catalog_sch_id(name, state.occupied_ids)
        await conn.execute(
            "UPDATE schools SET sch_id = $2, gaokao_school_id = $3 WHERE id = $1",
            existing_id,
            new_sch_id,
            gaokao_school_id,
        )
        state.occupied_ids.discard(sch_id)
        state.occupied_ids.add(new_sch_id)
        sch_id = new_sch_id
    else:
        await conn.execute(
            "UPDATE schools SET gaokao_school_id = $2 WHERE id = $1",
            existing_id,
            gaokao_school_id,
        )

    updated = {**dict(existing), "sch_id": sch_id, "gaokao_school_id": gaokao_school_id}
    state.existing_by_name[name] = updated
    state.existing_by_id[existing_id] = updated
    state.occupied_gaokao_ids[gaokao_school_id] = existing_id
    return "migrated" if legacy_placeholder else "linked"


async def _insert_gaokao_placeholder(
    conn: asyncpg.Connection,
    state: _GaokaoCatalogState,
    name: str,
    gaokao_school_id: int,
) -> str:
    if gaokao_school_id in state.occupied_gaokao_ids:
        return "conflicts"
    sch_id = _allocate_catalog_sch_id(name, state.occupied_ids)
    inserted = await conn.fetchrow(
        """
        INSERT INTO schools (sch_id, gaokao_school_id, name, is_211, is_985, is_double_first,
            is_private, is_independent, is_sino_foreign)
        VALUES ($1, $2, $3, FALSE, FALSE, FALSE, FALSE, FALSE, FALSE)
        ON CONFLICT DO NOTHING
        RETURNING id
        """,
        sch_id,
        gaokao_school_id,
        name,
    )
    if inserted is None:
        return "conflicts"
    school_id = int(inserted["id"])
    school = {
        "id": school_id,
        "name": name,
        "sch_id": sch_id,
        "gaokao_school_id": gaokao_school_id,
    }
    state.occupied_ids.add(sch_id)
    state.occupied_gaokao_ids[gaokao_school_id] = school_id
    state.existing_by_name[name] = school
    state.existing_by_id[school_id] = school
    return "added"


def _read_reference_schools(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        result: dict[str, dict[str, str]] = {}
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            name = str(row[7]).strip() if len(row) > 7 and row[7] else ""
            if name and name not in result:
                result[name] = {
                    "name": name,
                    "province": str(row[40]).strip() if len(row) > 40 and row[40] else "",
                    "city": str(row[41]).strip() if len(row) > 41 and row[41] else "",
                }
        return list(result.values())
    finally:
        workbook.close()


def _resolve_school(target: dict[str, str]) -> dict[str, object] | None:
    try:
        expected_name = _REFERENCE_ALIASES.get(target["name"], target["name"])
        html = _fetch(_SEARCH_URL.format(quote(expected_name)))
        candidates = _SCHOOL_ID_RE.findall(html)
        for sch_id in dict.fromkeys(candidates):
            payload = _fetch_json(_INFO_URL.format(sch_id))
            msg = payload.get("msg") if payload.get("flag") else None
            official_name = str(msg.get("yxmc", "")).strip() if isinstance(msg, dict) else ""
            if not isinstance(msg, dict) or official_name != expected_name:
                continue
            return {
                "sch_id": int(sch_id),
                "name": target["name"],
                "official_name": official_name,
                "province": target["province"],
                "province_id": None,
                "city": target["city"] or msg.get("yxszd"),
                "authority": msg.get("zgbmmc"),
                "level": None,
                "is_double_first": bool(msg.get("syl")),
            }
    except OSError:
        return None
    return None


def _safe_catalog_sch_id(school: dict[str, object]) -> int:
    sch_id = int(cast(int, school["sch_id"]))
    if 0 < sch_id <= 2_147_483_647 and school.get("official_name") == school.get("name"):
        return sch_id
    return _synthetic_catalog_sch_id(str(school["name"]))


def _allocate_catalog_sch_id(name: str, occupied_ids: set[int]) -> int:
    candidate = _synthetic_catalog_sch_id(name)
    while candidate in occupied_ids:
        candidate -= 1
    return candidate


def _school_row_priority(row: Any) -> tuple[bool, bool, bool, int]:
    sch_id = int(row["sch_id"])
    return (
        sch_id > 0 and row.get("crawl_task_id") is not None,
        sch_id > 0,
        row.get("gaokao_school_id") is not None,
        -int(row["id"]),
    )


def _is_legacy_gaokao_placeholder(row: Any, gaokao_school_id: int) -> bool:
    if row.get("gaokao_school_id") is not None or int(row["sch_id"]) != gaokao_school_id:
        return False
    data_fields = (
        "province_id",
        "city",
        "authority",
        "level",
        "school_type",
        "website",
        "phone",
        "email",
        "address",
        "introduction",
        "logo_url",
        "content_hash",
        "crawl_task_id",
    )
    flag_fields = (
        "is_211",
        "is_985",
        "is_double_first",
        "is_private",
        "is_independent",
        "is_sino_foreign",
    )
    return all(row.get(field) in (None, "") for field in data_fields) and not any(
        bool(row.get(field)) for field in flag_fields
    )


def _normalize_province_name(name: str) -> str:
    return (
        name
        .strip()
        .replace("壮族自治区", "")
        .replace("回族自治区", "")
        .replace("维吾尔自治区", "")
        .replace("自治区", "")
        .removesuffix("省")
        .removesuffix("市")
    )


def _synthetic_catalog_sch_id(name: str) -> int:
    digest = hashlib.sha256(name.encode("utf-8")).digest()
    return -(int.from_bytes(digest[:4], "big") % 2_000_000_000 + 1)


def _fetch(url: str) -> str:
    request = UrlRequest(url, headers={"User-Agent": "Mozilla/5.0"})  # noqa: S310
    with urlopen(request, timeout=20) as response:  # noqa: S310
        return response.read().decode("utf-8", errors="replace")


def _fetch_json(url: str) -> dict:
    try:
        return json.loads(_fetch(url))
    except (OSError, ValueError):
        return {}
