from __future__ import annotations

from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook

from gaokao_vault.exporters.liaoning import (
    build_liaoning_export_rows,
    history_years_for,
    merge_liaoning_baseline_rows,
    write_liaoning_workbook,
)


def _plan(**overrides) -> dict:
    plan = {
        "plan_id": 593,
        "year": 2026,
        "school_id": 1,
        "school_code_raw": "0001",
        "school_name": "测试大学",
        "school_province": "北京",
        "sch_id": 34,
        "city": "北京",
        "authority": "教育部",
        "school_level": "本科",
        "school_type": "综合",
        "school_crawl_task_id": 1,
        "is_211": True,
        "is_985": True,
        "is_double_first": True,
        "is_private": False,
        "is_independent": False,
        "is_sino_foreign": False,
        "major_id": 10,
        "major_code_raw": "01",
        "major_full_name": "计算机科学与技术(拔尖班)",
        "canonical_major_name": "计算机科学与技术",
        "major_category": "工学",
        "major_subcategory": "计算机类",
        "education_level": "本科",
        "subject_category": "物理类",
        "batch": "普通类本科批",
        "batch_code": "regular",
        "batch_category": "普通批",
        "batch_segment": None,
        "program_type": "普通类",
        "plan_count": 5,
        "duration": "四年",
        "tuition": "5200",
        "note": None,
        "selection_requirement": "物理+化学",
        "adjustment_rule": None,
        "major_strength_tier": "national",
        "strength_evidence": [
            {"signal_type": "first_class_major", "signal_level": "national"},
            {"signal_type": "featured_major", "signal_level": "national"},
        ],
    }
    plan.update(overrides)
    return plan


def _admission(**overrides) -> dict:
    admission = {
        "admission_id": 100,
        "school_id": 1,
        "major_id": 10,
        "year": 2025,
        "subject_category": "物理类",
        "batch": "本科批",
        "batch_code": "regular",
        "batch_category": "普通批",
        "batch_segment": None,
        "canonical_major_name": "计算机科学与技术",
        "major_name_raw": "计算机科学与技术",
        "admitted_count": 4,
        "min_score": 620,
        "min_rank": 3000,
        "plan_count": None,
        "data_source": "gaokao.chsi.com.cn",
    }
    admission.update(overrides)
    return admission


def test_build_liaoning_rows_matches_exact_history_and_keeps_leading_zeroes() -> None:
    plans = [_plan()]
    admissions = [
        _admission(),
        _admission(admission_id=101, year=2024, major_id=11),
        _admission(admission_id=102, year=2023, subject_category="历史类"),
        _admission(admission_id=103, year=2022, batch="本科提前批A段", batch_code="early"),
    ]
    historical_plans = [
        {
            "school_id": 1,
            "major_id": 10,
            "major_name": "计算机科学与技术",
            "canonical_major_name": "计算机科学与技术",
            "year": 2025,
            "subject_category": "物理类",
            "batch": "本科批",
            "batch_code": "regular",
            "batch_category": "普通批",
            "batch_segment": None,
            "plan_count": 5,
        }
    ]
    charters = [
        {
            "school_id": 1,
            "content": "对进档考生采用分数优先的方式安排专业。",
            "source_url": "https://example.com/charter",
        }
    ]

    rows, matched_counts, new_count = build_liaoning_export_rows(
        plans,
        admissions,
        historical_plans,
        charters,
        plan_year=2026,
    )

    assert len(rows) == 1
    row = rows[0]
    assert len(row) == 65
    assert row[0] == "593"
    assert row[1] == 2026
    assert row[6] == "0001"
    assert row[8] == "01"
    assert row[11] == "(拔尖班)"
    assert row[12] == "化学"
    assert row[15] == "4"
    assert row[19] is None
    assert row[20:25] == [4, 620, 3000, "本科批", 5]
    assert row[25:40] == [None] * 15
    assert row[43] == "985/211/双一流"
    assert row[48] == "公办"
    assert row[57] == "分数优先"
    assert row[58] == "https://example.com/charter"
    assert row[62] == "国一/国特"
    assert matched_counts == {2025: 1, 2024: 0, 2023: 0, 2022: 0}
    assert new_count == 0


def test_name_fallback_only_applies_when_plan_major_id_is_missing_and_unambiguous() -> None:
    plan = _plan(major_id=None)
    admission = _admission(major_id=20)

    rows, matched_counts, _new_count = build_liaoning_export_rows(
        [plan],
        [admission],
        [],
        [],
        plan_year=2026,
    )

    assert rows[0][20] == 4
    assert matched_counts[2025] == 1

    ambiguous = [admission, _admission(admission_id=101, major_id=21)]
    rows, matched_counts, new_count = build_liaoning_export_rows(
        [plan],
        ambiguous,
        [],
        [],
        plan_year=2026,
    )

    assert rows[0][20:40] == [None] * 20
    assert rows[0][19] == "新增"
    assert matched_counts[2025] == 0
    assert new_count == 1


def test_historical_plan_prevents_false_new_label_without_admission_result() -> None:
    historical_plan = {
        "school_id": 1,
        "major_id": 10,
        "major_name": "计算机科学与技术",
        "canonical_major_name": "计算机科学与技术",
        "year": 2025,
        "subject_category": "物理类",
        "batch": "本科批",
        "batch_code": "regular",
        "batch_category": "普通批",
        "batch_segment": None,
        "plan_count": 6,
    }

    rows, matched_counts, new_count = build_liaoning_export_rows(
        [_plan()],
        [],
        [historical_plan],
        [],
        plan_year=2026,
    )

    assert rows[0][19] is None
    assert rows[0][24] == 6
    assert matched_counts[2025] == 0
    assert new_count == 0


def test_merge_baseline_preserves_history_and_unseen_rows() -> None:
    generated_rows, _, _ = build_liaoning_export_rows(
        [_plan(plan_count=8)],
        [],
        [],
        [],
        plan_year=2026,
    )
    matched_baseline = list(generated_rows[0])
    matched_baseline[14] = 5
    matched_baseline[19] = None
    matched_baseline[20:25] = [4, 620, 3000, "本科批", 5]
    unseen_baseline = list(matched_baseline)
    unseen_baseline[0] = "999"
    unseen_baseline[7] = "未抓到大学"
    unseen_baseline[9] = "历史学"
    unseen_baseline[10] = "历史学"

    rows, matched_counts, new_count = merge_liaoning_baseline_rows(
        generated_rows,
        [matched_baseline, unseen_baseline],
        plan_year=2026,
    )

    assert len(rows) == 2
    assert rows[0][14] == 8
    assert rows[0][20:25] == [4, 620, 3000, "本科批", 5]
    assert rows[0][19] is None
    assert rows[1][7] == "未抓到大学"
    assert matched_counts[2025] == 2
    assert new_count == 0


def test_merge_baseline_matches_unique_name_when_source_codes_differ() -> None:
    generated_rows, _, _ = build_liaoning_export_rows(
        [_plan(school_code_raw=None, major_code_raw="080901")],
        [],
        [],
        [],
        plan_year=2026,
    )
    baseline = list(generated_rows[0])
    baseline[6] = "0001"
    baseline[8] = "48"
    baseline[20:25] = [4, 620, 3000, "本科批", 5]

    rows, _, _ = merge_liaoning_baseline_rows(generated_rows, [baseline], plan_year=2026)

    assert len(rows) == 1
    assert rows[0][8] == "080901"
    assert rows[0][20:25] == [4, 620, 3000, "本科批", 5]


def test_write_liaoning_workbook_matches_attachment_structure(tmp_path: Path) -> None:
    rows, _matched_counts, _new_count = build_liaoning_export_rows(
        [_plan()],
        [_admission(plan_count=5)],
        [],
        [],
        plan_year=2026,
    )
    output = tmp_path / "liaoning.xlsx"

    write_liaoning_workbook(output, rows, plan_year=2026)

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook.active
    assert worksheet.max_row == 3
    assert worksheet.max_column == 65
    assert worksheet.freeze_panes == "A3"
    assert worksheet.auto_filter.ref == "A2:BM3"
    assert worksheet.page_setup.orientation == "landscape"
    assert worksheet.page_setup.fitToWidth == 1
    assert history_years_for(2026) == (2025, 2024, 2023, 2022)
    assert worksheet["A1"].value == "2026年招生计划"
    assert worksheet["U1"].value == "2025年专业录取数据"
    assert worksheet["AJ1"].value == "2022年专业录取数据"
    assert worksheet["AO1"].value == "院校基础信息"
    assert worksheet["BH1"].value == "专业基础信息"
    assert worksheet["BM2"].value == "本专业博士点"
    assert {str(merged_range) for merged_range in worksheet.merged_cells.ranges} == {
        "A1:S1",
        "U1:Y1",
        "Z1:AD1",
        "AE1:AI1",
        "AJ1:AN1",
        "AO1:BG1",
        "BH1:BM1",
    }
    assert worksheet["A3"].value == "593"
    assert worksheet["B3"].value == 2026
    assert worksheet["G3"].value == "0001"
    assert worksheet["I3"].value == "01"
    assert worksheet["A3"].number_format == "@"
    assert worksheet["B3"].number_format != "@"
    assert worksheet["G3"].number_format == "@"
    assert worksheet["L2"].fill.fgColor.rgb == "0000B050"
    assert worksheet["O2"].fill.fgColor.rgb == "00FFFF00"
    assert worksheet["U2"].fill.fgColor.rgb == "00FFFF00"
    assert all(cell.data_type != "f" for row in worksheet.iter_rows() for cell in row)

    with ZipFile(output) as archive:
        styles = ElementTree.fromstring(archive.read("xl/styles.xml"))  # noqa: S314
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    expected_order = {
        name: index
        for index, name in enumerate((
            "b",
            "i",
            "strike",
            "outline",
            "shadow",
            "condense",
            "extend",
            "u",
            "vertAlign",
            "sz",
            "color",
            "name",
            "family",
            "charset",
            "scheme",
        ))
    }
    for font in styles.iter(f"{namespace}font"):
        positions = [expected_order[child.tag.rsplit("}", 1)[-1]] for child in font]
        assert positions == sorted(positions)


def test_history_matching_prefers_exact_raw_codes_for_same_canonical_major() -> None:
    plan = _plan(school_code_raw="0248", major_code_raw="04")
    admissions = [
        _admission(admission_id=101, school_code_raw="0248", major_code_raw="01", min_score=693),
        _admission(admission_id=102, school_code_raw="0248", major_code_raw="04", min_score=692),
    ]

    rows, matched_counts, _new_count = build_liaoning_export_rows(
        [plan],
        admissions,
        [],
        [],
        plan_year=2026,
    )

    assert rows[0][21] == 692
    assert matched_counts[2025] == 1


def test_history_matching_leaves_ambiguous_variants_blank() -> None:
    plan = _plan(school_code_raw=None, major_code_raw=None, major_full_name="计算机类")
    admissions = [
        _admission(admission_id=101, major_name_raw="计算机类(拔尖班)", major_code_raw="01"),
        _admission(admission_id=102, major_name_raw="计算机类(中外合作办学)", major_code_raw="02"),
    ]

    rows, matched_counts, new_count = build_liaoning_export_rows(
        [plan],
        admissions,
        [],
        [],
        plan_year=2026,
    )

    assert rows[0][20:40] == [None] * 20
    assert rows[0][19] == "新增"
    assert matched_counts[2025] == 0
    assert new_count == 1


def test_history_matching_survives_major_id_change_with_same_source_identity() -> None:
    plan = _plan(major_id=10, major_code_raw="04", major_full_name="计算机科学与技术(拔尖班)")
    admission = _admission(
        major_id=99,
        major_code_raw="04",
        major_name_raw="计算机科学与技术(拔尖班)",
        min_score=688,
    )

    rows, matched_counts, _new_count = build_liaoning_export_rows(
        [plan],
        [admission],
        [],
        [],
        plan_year=2026,
    )

    assert rows[0][21] == 688
    assert matched_counts[2025] == 1


def test_history_matching_does_not_treat_plain_major_as_named_variant() -> None:
    plan = _plan(major_code_raw=None, major_full_name="计算机科学与技术(中外合作办学)")
    admission = _admission(major_code_raw=None, major_name_raw="计算机科学与技术(普通班)")

    rows, matched_counts, new_count = build_liaoning_export_rows(
        [plan],
        [admission],
        [],
        [],
        plan_year=2026,
    )

    assert rows[0][20:25] == [None] * 5
    assert rows[0][19] == "新增"
    assert matched_counts[2025] == 0
    assert new_count == 1


def test_merge_baseline_subject_filter_drops_other_subject_rows() -> None:
    generated_rows, _, _ = build_liaoning_export_rows(
        [_plan(subject_category="物理类")],
        [],
        [],
        [],
        plan_year=2026,
    )
    physics = list(generated_rows[0])
    history = list(physics)
    history[0] = "history"
    history[4] = "历史"

    rows, _, _ = merge_liaoning_baseline_rows(
        generated_rows,
        [physics, history],
        plan_year=2026,
        subject="物理",
    )

    assert len(rows) == 1
    assert rows[0][4] == "物理"


def test_unknown_school_ownership_is_left_blank() -> None:
    plan = _plan(school_crawl_task_id=None, is_private=False, is_independent=False, is_sino_foreign=False)

    rows, _, _ = build_liaoning_export_rows([plan], [], [], [], plan_year=2026)

    assert rows[0][48] is None
